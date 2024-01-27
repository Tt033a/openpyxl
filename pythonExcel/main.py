import openpyxl
wb=openpyxl.Workbook()
ws=wb.active

ws.insert_rows(2)
for i in range(10):
    l=[]
ws.cell().value
ws.max_row